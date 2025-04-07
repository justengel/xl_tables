import os
import signal
from unittest import mock
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

from ..prop_utils import HashDict, ItemStorage
from ..fields import ConstantItem
from .fake_excel import Excel, fake_proxy_method, fake_proxy_property
from .csv_utils import csv_to_openpyxl, openpyxl_to_csv


def mock_borders():
    def get(self):
        borders = getattr(self, "_borders", None)
        if borders is None:
            self._borders = borders = mock.MagicMock()
        return borders
    
    return property(get)


class CellProxy:
    Borders = mock_borders()
    
    def __init__(self, cell):
        self._cell = cell

    @property
    def Value(self):
        return self._cell.value

    @Value.setter
    def Value(self, value):
        self._cell.value = value


class CellsCollection:
    Borders = mock_borders()

    def __init__(self, sheet, range_str):
        self.sheet = sheet
        self.range_str = range_str

    @property
    def Value(self):
        cells = self.sheet[self.range_str]
        try:
            len(cells)
            try:
                len(cells[0])
                return tuple(tuple(cell.value for cell in row) for row in cells)
            except TypeError:
                return tuple(cell.value for cell in cells)
        except TypeError:
            return cells.value
        
    @Value.setter
    def Value(self, values):
        cells = self.sheet[self.range_str]
        try:
            len(cells)
            try:
                len(cells[0])
                for i, row in enumerate(cells):
                    for j, cell in enumerate(row):
                        try:
                            cell.value = values[i][j]
                        except IndexError:
                            break
            except TypeError:
                for i, cell in enumerate(cells):
                    try:
                        cell.value = values[i]
                    except IndexError:
                        break
        except TypeError:
            cells.value = values

    def __repr__(self):
        return f"<Cells({self.range_str}, value={self.value})>"

    def __str__(self):
        return str(self.value)

    def __call__(self, row, column):
        """Access a cell by row and column (1-based indices)."""
        if self.range_str:
            return CellProxy(self.sheet[self.range_str][row-1][column-1])
        else:
            return CellProxy(self.sheet.cell(row=row, column=column))

    def __iter__(self):
        """Iterate over all cells."""
        return (CellProxy(cell) for row in self.sheet[self.range_str] for cell in row)


class RowsCollection:
    Borders = mock_borders()

    def __init__(self, sheet):
        self.sheet = sheet

    def __iter__(self):
        """Iterate over all rows."""
        return self.sheet.rows

    @property
    def count(self):
        """Number of rows."""
        return self.sheet.max_row

    @property
    def Value(self):
        """Get values of all rows."""
        return tuple(tuple(cell.value for cell in row) for row in self.sheet.rows)

    @Value.setter
    def Value(self, values):
        """Set values for all rows (assumes a 2D iterable)."""
        for i, row in enumerate(self.sheet.rows, 1):
            for j, cell in enumerate(row, 1):
                try:
                    cell.value = values[i-1][j-1]
                except IndexError:
                    break


class ColumnsCollection:
    Borders = mock_borders()

    def __init__(self, sheet):
        self.sheet = sheet

    def __iter__(self):
        """Iterate over all columns."""
        return self.sheet.columns

    @property
    def count(self):
        """Number of columns."""
        return self.sheet.max_column

    @property
    def Value(self):
        """Get values of all columns (transposed view)."""
        return tuple(tuple(cell.value for cell in col) for col in self.sheet.columns)

    @Value.setter
    def Value(self, values):
        """Set values for all columns (assumes a 2D iterable, column-wise)."""
        for j, col in enumerate(self.sheet.columns, 1):
            for i, cell in enumerate(col, 1):
                try:
                    cell.value = values[i-1][j-1]
                except IndexError:
                    break


class RangeObject:
    Borders = mock_borders()

    def __init__(self, sheet, range_str):
        self.sheet = sheet
        self.range_str = range_str
        self._areas = [area.strip() for area in self.range_str.split(',')]

    @property
    def Cells(self):
        """Return cells in the range as a tuple of tuples."""
        return CellsCollection(self.sheet, self.range_str)

    @property
    def Value(self):
        """Get values in the range."""
        return self.Cells.Value

    @Value.setter
    def Value(self, values):
        """Set values in the range."""
        self.Cells.Value = values

    @property
    def Areas(self):
        return AreasCollection(self)


class AreasCollection:
    Borders = mock_borders()

    def __init__(self, parent):
        self.parent = parent

    @property
    def Count(self):
        """Return the number of areas."""
        return len(self.parent._areas)

    def __iter__(self):
        """Iterate over areas, returning Range-like objects."""
        for area in self.parent._areas:
            yield RangeObject(self.parent.sheet, area)


class Sheet:
    Borders = mock_borders()

    def __init__(self, sheet):
        """Initialize with an openpyxl Worksheet object."""
        self.sheet = sheet

    @property
    def Cells(self):
        """Mimic Excel's Cells property: access by (row, col) or iterate all cells."""
        return CellsCollection(self.sheet, "")

    @property
    def Rows(self):
        """Mimic Excel's Rows property: iterate over all rows with value setting."""
        return RowsCollection(self.sheet)

    @property
    def Columns(self):
        """Mimic Excel's Columns property: iterate over all columns with value setting."""
        return ColumnsCollection(self.sheet)

    def Range(self, range_str):
        """Mimic Excel's Range property: access a range like 'A1:B2' with value setting."""
        return RangeObject(self.sheet, range_str)


class Workbook(object):
    METHOD_SETTER_ERROR = 'Cannot set property '
    SAVE_ON_CLOSE = False

    def __init__(self, filename=None, *args, xl=None, wb=None, **xl_settings):
        # Variables
        self._xl = xl
        self._wb = wb
        self._filename = None  # Save the filename as a variable

        # Initialize Excel
        if self._xl is None:
            self._xl = Excel(**xl_settings)

        # Set the filename
        self.set_filename(filename)

        # Check to open the filename
        if isinstance(self.filename, str) and os.path.exists(self.filename) and os.path.isfile(self.filename):
            self.open(filename)

        # Initialize constants
        self.init_constants()

    def init_constants(self):
        """Set all of the constant values."""
        for k, field in self.__class__.__dict__.items():
            if isinstance(field, ConstantItem):
                field.init_table(self)

    @property
    def xl(self):
        """Get (or create) the Excel Application object."""
        if self._xl is None:
            self._xl = Excel()
        return self._xl

    @xl.setter
    def xl(self, value):
        """Set the Excel Application object."""
        self._xl = value

    @property
    def wb(self):
        """Get (or Add) a Workbook to the Excel Application Workbooks collection."""
        if self._wb is None:
            self._wb = openpyxl.Workbook()
        return self._wb

    @wb.setter
    def wb(self, value):
        """Set the Workbook object."""
        self._wb = value

    def get_filename(self):
        """Return the filename."""
        return self._filename

    def set_filename(self, filename):
        """Set the filename.

        Args:
            filename (str/object): Filename to save and open from.
        """
        if isinstance(filename, str):
            filename = os.path.abspath(filename)
        self._filename = filename

    filename = property(get_filename, set_filename)

    VALID_FMT = [".xlsx", ".xlsm", ".xltx", ".xltm"]

    def open(self, filename=None):
        """Open a workbook with the given filename and use this workbook."""
        if filename is not None:
            self.set_filename(filename)

        filename = self.get_filename()
        if isinstance(filename, str) and os.path.exists(filename) and os.path.isfile(filename):
            if os.path.splitext(filename.lower())[-1] in self.VALID_FMT:
                self._wb = openpyxl.load_workbook(filename)
            else:
                self._wb = csv_to_openpyxl(filename)
        return self
    
    def save(self, filename=None):
        """Save the given filename or set filename."""
        if filename is not None:
            self.set_filename(filename)

        filename = self.get_filename()

        # Saving as CSV or non excel type renames the active sheet to the base filename.
        if os.path.splitext(filename.lower())[-1] in self.VALID_FMT:
            self.wb.save(filename)
        else:
            openpyxl_to_csv(filename, self.wb)

    def get_sheet(self, sheet, create=True):
        """Return the sheet for an index or name."""
        try:
            if isinstance(sheet, int):
                obj = self.wb.worksheets[sheet-1]  # Get the sheet
            else:
                obj = self.wb[sheet]  # Get the sheet
            self.wb.active = self.wb.worksheets.index(obj)
        except (ValueError, TypeError, Exception):
            if create:
                if isinstance(sheet, int):
                    sheet = f"Sheet{sheet}"
                obj = self.wb.create_sheet(sheet)  # Create the sheet
                self.wb.active = obj
            else:
                obj = None
        return Sheet(obj)

    def has_sheet(self, sheet):
        """Return if the given sheet name or index exists"""
        return self.get_sheet(sheet, create=False) is not None
    
    @property
    def Cells(self):
        return Sheet(self.wb.active).Cells
    
    @property
    def Columns(self):
        return Sheet(self.wb.active).Columns
    
    @property
    def Rows(self):
        return Sheet(self.wb.active).Rows
    
    @property
    def Range(self):
        return Sheet(self.wb.active).Range
    
    # ===== Workbook Object Methods =====
    AcceptAllChanges = fake_proxy_method()
    Activate = fake_proxy_method()
    AddToFavorites = fake_proxy_method()
    ApplyTheme = fake_proxy_method()
    BreakLink = fake_proxy_method()
    CanCheckIn = fake_proxy_method()
    ChangeFileAccess = fake_proxy_method()
    ChangeLink = fake_proxy_method()
    CheckIn = fake_proxy_method()
    CheckInWithVersion = fake_proxy_method()
    Close = fake_proxy_method()
    ConvertComments = fake_proxy_method()
    CreateForecastSheet = fake_proxy_method()
    DeleteNumberFormat = fake_proxy_method()
    EnableConnections = fake_proxy_method()
    EndReview = fake_proxy_method()
    ExclusiveAccess = fake_proxy_method()
    ExportAsFixedFormat = fake_proxy_method()
    FollowHyperlink = fake_proxy_method()
    ForwardMailer = fake_proxy_method()
    GetWorkflowTasks = fake_proxy_method()
    GetWorkflowTemplates = fake_proxy_method()
    HighlightChangesOptions = fake_proxy_method()
    LinkInfo = fake_proxy_method()
    LinkSources = fake_proxy_method()
    LockServerFile = fake_proxy_method()
    MergeWorkbook = fake_proxy_method()
    NewWindow = fake_proxy_method()
    OpenLinks = fake_proxy_method()
    PivotCaches = fake_proxy_method()
    Post = fake_proxy_method()
    PrintOut = fake_proxy_method()
    PrintPreview = fake_proxy_method()
    Protect = fake_proxy_method()
    ProtectSharing = fake_proxy_method()
    PublishToDocs = fake_proxy_method()
    PurgeChangeHistoryNow = fake_proxy_method()
    RefreshAll = fake_proxy_method()
    RejectAllChanges = fake_proxy_method()
    ReloadAs = fake_proxy_method()
    RemoveDocumentInformation = fake_proxy_method()
    RemoveUser = fake_proxy_method()
    Reply = fake_proxy_method()
    ReplyAll = fake_proxy_method()
    ReplyWithChanges = fake_proxy_method()
    ResetColors = fake_proxy_method()
    RunAutoMacros = fake_proxy_method()
    Save = fake_proxy_method()
    SaveAs = fake_proxy_method()
    SaveAsXMLData = fake_proxy_method()
    SaveCopyAs = fake_proxy_method()
    SendFaxOverInternet = fake_proxy_method()
    SendForReview = fake_proxy_method()
    SendMail = fake_proxy_method()
    SendMailer = fake_proxy_method()
    SetLinkOnData = fake_proxy_method()
    SetPasswordEncryptionOptions = fake_proxy_method()
    ToggleFormsDesign = fake_proxy_method()
    Unprotect = fake_proxy_method()
    UnprotectSharing = fake_proxy_method()
    UpdateFromFile = fake_proxy_method()
    UpdateLink = fake_proxy_method()
    WebPagePreview = fake_proxy_method()
    XmlImport = fake_proxy_method()
    XmlImportXml = fake_proxy_method()

    # ===== Workbook Object Properties =====
    AccuracyVersion = fake_proxy_property()
    ActiveChart = fake_proxy_property()
    ActiveSheet = fake_proxy_property()
    ActiveSlicer = fake_proxy_property()
    Application = fake_proxy_property()
    AutoSaveOn = fake_proxy_property()
    AutoUpdateFrequency = fake_proxy_property()
    AutoUpdateSaveChanges = fake_proxy_property()
    BuiltinDocumentProperties = fake_proxy_property()
    CalculationVersion = fake_proxy_property()
    CaseSensitive = fake_proxy_property()
    ChangeHistoryDuration = fake_proxy_property()
    ChartDataPointTrack = fake_proxy_property()
    Charts = fake_proxy_property()
    CheckCompatibility = fake_proxy_property()
    CodeName = fake_proxy_property()
    Colors = fake_proxy_property()
    CommandBars = fake_proxy_property()
    ConflictResolution = fake_proxy_property()
    Connections = fake_proxy_property()
    ConnectionsDisabled = fake_proxy_property()
    Container = fake_proxy_property()
    ContentTypeProperties = fake_proxy_property()
    CreateBackup = fake_proxy_property()
    Creator = fake_proxy_property()
    CustomDocumentProperties = fake_proxy_property()
    CustomViews = fake_proxy_property()
    CustomXMLParts = fake_proxy_property()
    Date1904 = fake_proxy_property()
    DefaultPivotTableStyle = fake_proxy_property()
    DefaultSlicerStyle = fake_proxy_property()
    DefaultTableStyle = fake_proxy_property()
    DefaultTimelineStyle = fake_proxy_property()
    DisplayDrawingObjects = fake_proxy_property()
    DisplayInkComments = fake_proxy_property()
    DocumentInspectors = fake_proxy_property()
    DocumentLibraryVersions = fake_proxy_property()
    DoNotPromptForConvert = fake_proxy_property()
    EnableAutoRecover = fake_proxy_property()
    EncryptionProvider = fake_proxy_property()
    EnvelopeVisible = fake_proxy_property()
    Excel4IntlMacroSheets = fake_proxy_property()
    Excel4MacroSheets = fake_proxy_property()
    Excel8CompatibilityMode = fake_proxy_property()
    FileFormat = fake_proxy_property()
    Final = fake_proxy_property()
    ForceFullCalculation = fake_proxy_property()
    FullName = fake_proxy_property()
    FullNameURLEncoded = fake_proxy_property()
    HasPassword = fake_proxy_property()
    HasVBProject = fake_proxy_property()
    HighlightChangesOnScreen = fake_proxy_property()
    IconSets = fake_proxy_property()
    InactiveListBorderVisible = fake_proxy_property()
    IsAddin = fake_proxy_property()
    IsInplace = fake_proxy_property()
    KeepChangeHistory = fake_proxy_property()
    ListChangesOnNewSheet = fake_proxy_property()
    Mailer = fake_proxy_property()
    Model = fake_proxy_property()
    MultiUserEditing = fake_proxy_property()
    Name = fake_proxy_property()
    Names = fake_proxy_property()
    Parent = fake_proxy_property()
    Password = fake_proxy_property()
    PasswordEncryptionAlgorithm = fake_proxy_property()
    PasswordEncryptionFileProperties = fake_proxy_property()
    PasswordEncryptionKeyLength = fake_proxy_property()
    PasswordEncryptionProvider = fake_proxy_property()
    Path = fake_proxy_property()
    Permission = fake_proxy_property()
    PersonalViewListSettings = fake_proxy_property()
    PersonalViewPrintSettings = fake_proxy_property()
    PivotTables = fake_proxy_property()
    PrecisionAsDisplayed = fake_proxy_property()
    ProtectStructure = fake_proxy_property()
    ProtectWindows = fake_proxy_property()
    PublishObjects = fake_proxy_property()
    Queries = fake_proxy_property()
    ReadOnly = fake_proxy_property()
    ReadOnlyRecommended = fake_proxy_property()
    RemovePersonalInformation = fake_proxy_property()
    Research = fake_proxy_property()
    RevisionNumber = fake_proxy_property()
    Saved = fake_proxy_property()
    SaveLinkValues = fake_proxy_property()
    ServerPolicy = fake_proxy_property()
    ServerViewableItems = fake_proxy_property()
    SharedWorkspace = fake_proxy_property()
    # Sheets = fake_proxy_property()
    ShowConflictHistory = fake_proxy_property()
    ShowPivotChartActiveFields = fake_proxy_property()
    ShowPivotTableFieldList = fake_proxy_property()
    Signatures = fake_proxy_property()
    SlicerCaches = fake_proxy_property()
    SmartDocument = fake_proxy_property()
    Styles = fake_proxy_property()
    Sync = fake_proxy_property()
    TableStyles = fake_proxy_property()
    TemplateRemoveExtData = fake_proxy_property()
    Theme = fake_proxy_property()
    UpdateLinks = fake_proxy_property()
    UpdateRemoteReferences = fake_proxy_property()
    UserStatus = fake_proxy_property()
    UseWholeCellCriteria = fake_proxy_property()
    UseWildcards = fake_proxy_property()
    VBASigned = fake_proxy_property()
    VBProject = fake_proxy_property()
    WebOptions = fake_proxy_property()
    Windows = fake_proxy_property()
    Worksheets = fake_proxy_property()
    WritePassword = fake_proxy_property()
    WriteReserved = fake_proxy_property()
    WriteReservedBy = fake_proxy_property()
    XmlMaps = fake_proxy_property()
    XmlNamespaces = fake_proxy_property()

    # ===== Application Methods =====
    ActivateMicrosoftApp = fake_proxy_method()
    AddCustomList = fake_proxy_method()
    Calculate = fake_proxy_method()
    CalculateFull = fake_proxy_method()
    CalculateFullRebuild = fake_proxy_method()
    CalculateUntilAsyncQueriesDone = fake_proxy_method()
    CentimetersToPoints = fake_proxy_method()
    CheckAbort = fake_proxy_method()
    CheckSpelling = fake_proxy_method()
    ConvertFormula = fake_proxy_method()
    DDEExecute = fake_proxy_method()
    DDEInitiate = fake_proxy_method()
    DDEPoke = fake_proxy_method()
    DDERequest = fake_proxy_method()
    DDETerminate = fake_proxy_method()
    DeleteCustomList = fake_proxy_method()
    DisplayXMLSourcePane = fake_proxy_method()
    DoubleClick = fake_proxy_method()
    Evaluate = fake_proxy_method()
    ExecuteExcel4Macro = fake_proxy_method()
    FindFile = fake_proxy_method()
    GetCustomListContents = fake_proxy_method()
    GetCustomListNum = fake_proxy_method()
    GetOpenFilename = fake_proxy_method()
    GetPhonetic = fake_proxy_method()
    GetSaveAsFilename = fake_proxy_method()
    Goto = fake_proxy_method()
    Help = fake_proxy_method()
    InchesToPoints = fake_proxy_method()
    InputBox = fake_proxy_method()
    Intersect = fake_proxy_method()
    MacroOptions = fake_proxy_method()
    MailLogoff = fake_proxy_method()
    MailLogon = fake_proxy_method()
    NextLetter = fake_proxy_method()
    OnKey = fake_proxy_method()
    OnRepeat = fake_proxy_method()
    OnTime = fake_proxy_method()
    OnUndo = fake_proxy_method()
    Quit = fake_proxy_method()
    RecordMacro = fake_proxy_method()
    RegisterXLL = fake_proxy_method()
    Repeat = fake_proxy_method()
    Run = fake_proxy_method()
    SendKeys = fake_proxy_method()
    SharePointVersion = fake_proxy_method()
    Undo = fake_proxy_method()
    Union = fake_proxy_method()
    Volatile = fake_proxy_method()
    Wait = fake_proxy_method()

    # ===== Application Properties =====
    ActiveCell = fake_proxy_property()
    # ActiveChart = fake_proxy_property()
    ActiveEncryptionSession = fake_proxy_property()
    ActivePrinter = fake_proxy_property()
    ActiveProtectedViewWindow = fake_proxy_property()
    # ActiveSheet = fake_proxy_property()
    ActiveWindow = fake_proxy_property()
    ActiveWorkbook = fake_proxy_property()
    AddIns = fake_proxy_property()
    AddIns2 = fake_proxy_property()
    AlertBeforeOverwriting = fake_proxy_property()
    AltStartupPath = fake_proxy_property()
    AlwaysUseClearType = fake_proxy_property()
    # Application = fake_proxy_property()
    ArbitraryXMLSupportAvailable = fake_proxy_property()
    AskToUpdateLinks = fake_proxy_property()
    Assistance = fake_proxy_property()
    AutoCorrect = fake_proxy_property()
    AutoFormatAsYouTypeReplaceHyperlinks = fake_proxy_property()
    AutomationSecurity = fake_proxy_property()
    AutoPercentEntry = fake_proxy_property()
    AutoRecover = fake_proxy_property()
    Build = fake_proxy_property()
    CalculateBeforeSave = fake_proxy_property()
    Calculation = fake_proxy_property()
    CalculationInterruptKey = fake_proxy_property()
    CalculationState = fake_proxy_property()
    # CalculationVersion = fake_proxy_property()
    Caller = fake_proxy_property()
    CanPlaySounds = fake_proxy_property()
    CanRecordSounds = fake_proxy_property()
    Caption = fake_proxy_property()
    CellDragAndDrop = fake_proxy_property()
    #>>>> Cells = fake_proxy_property()
    # ChartDataPointTrack = fake_proxy_property()
    # Charts = fake_proxy_property()
    ClipboardFormats = fake_proxy_property()
    ClusterConnector = fake_proxy_property()
    #>>>> Columns = fake_proxy_property()
    COMAddIns = fake_proxy_property()
    # CommandBars = fake_proxy_property()
    CommandUnderlines = fake_proxy_property()
    ConstrainNumeric = fake_proxy_property()
    ControlCharacters = fake_proxy_property()
    CopyObjectsWithCells = fake_proxy_property()
    # Creator = fake_proxy_property()
    Cursor = fake_proxy_property()
    CursorMovement = fake_proxy_property()
    CustomListCount = fake_proxy_property()
    CutCopyMode = fake_proxy_property()
    DataEntryMode = fake_proxy_property()
    DDEAppReturnCode = fake_proxy_property()
    DecimalSeparator = fake_proxy_property()
    DefaultFilePath = fake_proxy_property()
    DefaultSaveFormat = fake_proxy_property()
    DefaultSheetDirection = fake_proxy_property()
    DefaultWebOptions = fake_proxy_property()
    DeferAsyncQueries = fake_proxy_property()
    Dialogs = fake_proxy_property()
    DisplayAlerts = fake_proxy_property()
    DisplayClipboardWindow = fake_proxy_property()
    DisplayCommentIndicator = fake_proxy_property()
    DisplayDocumentActionTaskPane = fake_proxy_property()
    DisplayDocumentInformationPanel = fake_proxy_property()
    DisplayExcel4Menus = fake_proxy_property()
    DisplayFormulaAutoComplete = fake_proxy_property()
    DisplayFormulaBar = fake_proxy_property()
    DisplayFullScreen = fake_proxy_property()
    DisplayFunctionToolTips = fake_proxy_property()
    DisplayInsertOptions = fake_proxy_property()
    DisplayNoteIndicator = fake_proxy_property()
    DisplayPasteOptions = fake_proxy_property()
    DisplayRecentFiles = fake_proxy_property()
    DisplayScrollBars = fake_proxy_property()
    DisplayStatusBar = fake_proxy_property()
    EditDirectlyInCell = fake_proxy_property()
    EnableAnimations = fake_proxy_property()
    EnableAutoComplete = fake_proxy_property()
    EnableCancelKey = fake_proxy_property()
    EnableCheckFileExtensions = fake_proxy_property()
    EnableEvents = fake_proxy_property()
    EnableLargeOperationAlert = fake_proxy_property()
    EnableLivePreview = fake_proxy_property()
    EnableMacroAnimations = fake_proxy_property()
    EnableSound = fake_proxy_property()
    ErrorCheckingOptions = fake_proxy_property()
    # Excel4IntlMacroSheets = fake_proxy_property()
    # Excel4MacroSheets = fake_proxy_property()
    ExtendList = fake_proxy_property()
    FeatureInstall = fake_proxy_property()
    FileConverters = fake_proxy_property()
    FileDialog = fake_proxy_property()
    FileExportConverters = fake_proxy_property()
    FileValidation = fake_proxy_property()
    FileValidationPivot = fake_proxy_property()
    FindFormat = fake_proxy_property()
    FixedDecimal = fake_proxy_property()
    FixedDecimalPlaces = fake_proxy_property()
    FlashFill = fake_proxy_property()
    FlashFillMode = fake_proxy_property()
    FormulaBarHeight = fake_proxy_property()
    GenerateGetPivotData = fake_proxy_property()
    GenerateTableRefs = fake_proxy_property()
    Height = fake_proxy_property()
    HighQualityModeForGraphics = fake_proxy_property()
    Hinstance = fake_proxy_property()
    HinstancePtr = fake_proxy_property()
    Hwnd = fake_proxy_property()
    IgnoreRemoteRequests = fake_proxy_property()
    Interactive = fake_proxy_property()
    International = fake_proxy_property()
    IsSandboxed = fake_proxy_property()
    Iteration = fake_proxy_property()
    LanguageSettings = fake_proxy_property()
    LargeOperationCellThousandCount = fake_proxy_property()
    Left = fake_proxy_property()
    LibraryPath = fake_proxy_property()
    MailSession = fake_proxy_property()
    MailSystem = fake_proxy_property()
    MapPaperSize = fake_proxy_property()
    MathCoprocessorAvailable = fake_proxy_property()
    MaxChange = fake_proxy_property()
    MaxIterations = fake_proxy_property()
    MeasurementUnit = fake_proxy_property()
    MergeInstances = fake_proxy_property()
    MouseAvailable = fake_proxy_property()
    MoveAfterReturn = fake_proxy_property()
    MoveAfterReturnDirection = fake_proxy_property()
    MultiThreadedCalculation = fake_proxy_property()
    # Name = fake_proxy_property()
    # Names = fake_proxy_property()
    NetworkTemplatesPath = fake_proxy_property()
    NewWorkbook = fake_proxy_property()
    ODBCErrors = fake_proxy_property()
    ODBCTimeout = fake_proxy_property()
    OLEDBErrors = fake_proxy_property()
    OnWindow = fake_proxy_property()
    OperatingSystem = fake_proxy_property()
    OrganizationName = fake_proxy_property()
    # Parent = fake_proxy_property()
    # Path = fake_proxy_property()
    PathSeparator = fake_proxy_property()
    PivotTableSelection = fake_proxy_property()
    PreviousSelections = fake_proxy_property()
    PrintCommunication = fake_proxy_property()
    ProductCode = fake_proxy_property()
    PromptForSummaryInfo = fake_proxy_property()
    ProtectedViewWindows = fake_proxy_property()
    QuickAnalysis = fake_proxy_property()
    #>>>> Range = fake_proxy_property()
    Ready = fake_proxy_property()
    RecentFiles = fake_proxy_property()
    RecordRelative = fake_proxy_property()
    ReferenceStyle = fake_proxy_property()
    RegisteredFunctions = fake_proxy_property()
    ReplaceFormat = fake_proxy_property()
    RollZoom = fake_proxy_property()
    #>>>> Rows = fake_proxy_property()
    RTD = fake_proxy_property()
    ScreenUpdating = fake_proxy_property()
    Selection = fake_proxy_property()
    # Sheets = fake_proxy_property()
    SheetsInNewWorkbook = fake_proxy_property()
    ShowChartTipNames = fake_proxy_property()
    ShowChartTipValues = fake_proxy_property()
    ShowDevTools = fake_proxy_property()
    ShowMenuFloaties = fake_proxy_property()
    ShowQuickAnalysis = fake_proxy_property()
    ShowSelectionFloaties = fake_proxy_property()
    ShowStartupDialog = fake_proxy_property()
    ShowToolTips = fake_proxy_property()
    SmartArtColors = fake_proxy_property()
    SmartArtLayouts = fake_proxy_property()
    SmartArtQuickStyles = fake_proxy_property()
    Speech = fake_proxy_property()
    SpellingOptions = fake_proxy_property()
    StandardFont = fake_proxy_property()
    StandardFontSize = fake_proxy_property()
    StartupPath = fake_proxy_property()
    StatusBar = fake_proxy_property()
    TemplatesPath = fake_proxy_property()
    ThisCell = fake_proxy_property()
    ThisWorkbook = fake_proxy_property()
    ThousandsSeparator = fake_proxy_property()
    Top = fake_proxy_property()
    TransitionMenuKey = fake_proxy_property()
    TransitionMenuKeyAction = fake_proxy_property()
    TransitionNavigKeys = fake_proxy_property()
    UsableHeight = fake_proxy_property()
    UsableWidth = fake_proxy_property()
    UseClusterConnector = fake_proxy_property()
    UsedObjects = fake_proxy_property()
    UserControl = fake_proxy_property()
    UserLibraryPath = fake_proxy_property()
    UserName = fake_proxy_property()
    UseSystemSeparators = fake_proxy_property()
    Value = fake_proxy_property()
    VBE = fake_proxy_property()
    Version = fake_proxy_property()
    Visible = fake_proxy_property()
    WarnOnFunctionNameConflict = fake_proxy_property()
    Watches = fake_proxy_property()
    Width = fake_proxy_property()
    # Windows = fake_proxy_property()
    WindowsForPens = fake_proxy_property()
    WindowState = fake_proxy_property()
    Workbooks = fake_proxy_property()
    WorksheetFunction = fake_proxy_property()
    # Worksheets = fake_proxy_property()


def should_init_sig(*args, **kwargs):
    pass


def set_init_sig(*args, **kwargs):
    pass


def init_sig_shutdown(func=None):
    """Set the SIGTERM and SIGINT handlers to call "shutdown" and close all Excel Applications run by this process."""
    if func is None:
        func = shutdown
    signal.signal(signal.SIGTERM, func)
    signal.signal(signal.SIGINT, func)


def shutdown(*args, **kwargs):
    pass