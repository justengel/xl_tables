import csv
import openpyxl


def csv_to_openpyxl(csv_path):
    """
    Load a CSV file into a new openpyxl Workbook and save it as an Excel file.
    
    Args:
        csv_path (str): Path to the input CSV file.

    Returns:
        Workbook
    """
    # Create a new workbook
    wb = openpyxl.Workbook()
    ws = wb.active  # Use the default active sheet
    ws.title = "Sheet1"

    # Open and read the CSV file
    with open(csv_path, newline='', encoding='utf-8') as csvfile:
        reader = csv.reader(csvfile)
        # Iterate over each row in the CSV
        for row_idx, row in enumerate(reader, 1):  # 1-based indexing for openpyxl
            for col_idx, value in enumerate(row, 1):
                # Write each value to the corresponding cell
                ws.cell(row=row_idx, column=col_idx, value=value)

    return wb


def openpyxl_to_csv(csv_path, wb):
    with open(csv_path, 'w', newline='') as csvfile:
        writer = csv.writer(csvfile)
        for i, ws in enumerate(wb.worksheets):
            if len(wb.worksheets) > 1:
                if i > 0:
                    writer.writerow([])
                    writer.writerow([])
                writer.writerow([f"# {ws.title}"])

            for row in ws.iter_rows(values_only=True):
                writer.writerow(row)